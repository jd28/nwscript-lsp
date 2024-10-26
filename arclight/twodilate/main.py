#!/usr/bin/env python

import argparse
import csv
import ctypes
import glob
import importlib.resources
import io
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import os
from pathlib import Path
from prettytable import PrettyTable, PLAIN_COLUMNS
import re
import rollnw
import time
import yaml
import zipfile

from typing import Optional, List


def quote(string: str):
    return '"' + string + '"' if ' ' in string else string


def twoda_to_excel(twoda: rollnw.TwoDA, file_path: str):
    wb = Workbook()
    ws = wb.active

    ws.cell(row=1, column=1, value="2DA V2.0")

    for col_index, column_name in enumerate(twoda.column_names(), start=2):
        ws.cell(row=3, column=col_index, value=column_name)
        column_letter = get_column_letter(col_index)
        current_width = ws.column_dimensions[column_letter].width or 0
        new_width = max(current_width, len(column_name) + 8)
        ws.column_dimensions[column_letter].width = new_width

    for i in range(twoda.rows()):
        ws.cell(row=i+4, column=1, value=str(i))
        for j in range(twoda.columns()):
            value = twoda.get_raw(i, j)
            ws.cell(row=i+4, column=j+2, value=value)

            column_letter = get_column_letter(j + 2)
            current_width = ws.column_dimensions[column_letter].width or 0
            new_width = max(current_width, len(value) + 8)
            ws.column_dimensions[column_letter].width = new_width

    wb.save(file_path)


def twoda_from_excel(sheet) -> rollnw.TwoDA:
    twoda = rollnw.TwoDA()

    for col in range(2, sheet.max_column + 1):  # 1-based index
        cn = sheet.cell(row=3, column=col).value
        twoda.add_column(cn)

    # Adjust for excel row start, will segfault below if wrong
    twoda.pad(sheet.max_row - 3)

    for row in range(4, sheet.max_row + 1):
        for col in range(2, sheet.max_column + 1):
            r = row - 4
            c = col - 2
            cell_value = sheet.cell(row=row, column=col).value
            twoda.set(r, c, cell_value)

    return twoda


class TwoDX:
    """2dx Files.
    """

    def __init__(self, source: str):
        self.columns: List[str] = []
        self.rows: List[List[str]] = []
        self.max = None
        self.newline: str = "\n"
        self.metadata = {}
        self.metadata_str: str = ""
        if (len(source)):
            self.parse(source)

    def __getitem__(self, i):
        if isinstance(i, int):
            if i >= len(self.rows) or i < 0:
                raise ValueError("Invalid row index!")
            return self.rows[i]
        elif isinstance(i, slice):
            pass

    def __str__(self):
        result = io.StringIO()
        result.write("2DX V2.1\n")
        result.write("---\n")
        result.write(yaml.dump(self.metadata))
        result.write("---\n")

        table = PrettyTable(self.columns)
        table.set_style(PLAIN_COLUMNS)
        table.align = 'l'
        table.padding_width = 0

        for rs in self.rows:
            table.add_row([quote(word) for word in rs])

        for line in table.get_string().splitlines():
            result.write(line.rstrip())
            result.write("\n")
        return result.getvalue()

    def get(self, row, col):
        """Gets a 2dx entry by row and column label or column index.
        """
        col = self.column_index(col)
        return self.rows[row][col]

    def column_index(self, col: str):
        """Gets the column index from a column label.
        """
        return self.columns.index(col)

    def get_float(self, row, col):
        """Gets a 2dx entry by row and column label or column index as a float.
        """
        return float(self.get(row, col))

    def get_int(self, row, col):
        """Gets a 2dx entry by row and column label or column index as an int.
        """
        return int(self.get(row, col))

    def parse21Header(self, lines):
        i = 1
        holder = []

        while len(lines[i].strip()) == 0:
            i += 1

        if lines[i].startswith("---"):
            i += 1
            while not lines[i].startswith("---"):
                holder.append(lines[i])
                i += 1
                if i >= len(lines):
                    raise RuntimeError("Unterminated YAML header!")
            i += 1
            holder = '\n'.join(holder)
            self.metadata_str = holder
            self.metadata = yaml.load(holder, Loader=yaml.SafeLoader)

        return i

    def parse(self, content: str):
        """Parses a 2dx file.
        """
        lines = [l for l in iter(content.splitlines())]
        if len(lines) == 0:
            raise ValueError("Invalid 2dx file!")

        if re.match(r"2DX\s+V2.1", lines[0]):
            col_line = self.parse21Header(lines)
            self.version = lines[0]
        else:
            raise ValueError("Invalid 2DX header!")

        lines = [l.strip() for l in lines[col_line:] if len(l.strip()) > 0]
        csvreader = csv.reader(lines, delimiter=' ', skipinitialspace=True)
        for row in csvreader:
            self.rows.append(row)

        # 2dx doesn't need to have any rows/labels.  All changes can be in the metadata.
        if not len(self.rows):
            return

        self.columns = [''] + self.rows[0]
        self.rows = self.rows[1:]

    def set(self, row, col, val):
        """Sets a 2dx entry by row and column label or column index.
        """
        col = self.column_index(col)
        self.rows[row][col] = val

    def update_rows(self):
        if 'row' in self.metadata:
            row = self.metadata['row']
            for i in range(len(self.rows)):
                cur = self.rows[i][0]
                if cur == '****':
                    self.rows[i][0] = str(row + i)

    def update_tlks(self):
        if 'tlk' in self.metadata:
            for c, off in self.metadata['tlk'].items():
                for i in range(len(self.rows)):
                    cur = self.get(i, c)
                    if cur != '****':
                        self.set(i, c, str(int(cur) + int(off) + 0x01000000))

    def to_excel(self, file_path):
        """Exports the 2dx data to an Excel file."""
        wb = Workbook()
        ws = wb.active

        ws.cell(row=1, column=1, value="2DX V2.1")
        ws.cell(row=2, column=1, value=yaml.dump(self.metadata))

        for col_index, column_name in enumerate(self.columns, start=1):
            ws.cell(row=3, column=col_index, value=column_name)
            column_letter = get_column_letter(col_index)
            current_width = ws.column_dimensions[column_letter].width or 0
            new_width = max(current_width, len(column_name) + 8)
            ws.column_dimensions[column_letter].width = new_width

        for row_index, row in enumerate(self.rows, start=4):
            for col_index, cell_value in enumerate(row, start=1):
                ws.cell(row=row_index, column=col_index, value=cell_value)
                column_letter = get_column_letter(col_index)
                current_width = ws.column_dimensions[column_letter].width or 0
                new_width = max(current_width, len(cell_value) + 8)
                ws.column_dimensions[column_letter].width = new_width

        wb.save(file_path)

    def from_excel(self, ws):
        """Imports 2dx data from an Excel file."""

        self.columns = []
        self.rows = []
        header = ws.cell(row=1, column=1).value
        if header != "2DX V2.1":
            raise ValueError("Invalid file format. Missing header.")

        self.metadata_str = ws.cell(row=2, column=1).value
        self.metadata = yaml.load(self.metadata_str, Loader=yaml.SafeLoader)
        self.columns = [''] + [cell.value for cell in ws[3]
                               if cell.value is not None]
        for row in ws.iter_rows(min_row=4, values_only=True):
            self.rows.append([cell for cell in row if cell is not None])


class TwoDXMerger:
    def __init__(self, twoda: rollnw.TwoDA, twodx: TwoDX, default: Optional[rollnw.StaticTwoDA] = None):
        self.twoda: rollnw.TwoDA = twoda
        self.twodx: TwoDX = twodx
        self.default: rollnw.StaticTwoDA = default

    def merge(self):
        highest = 0
        nrows = self.twoda.rows()
        for r in self.twodx.rows:
            if r[0] == "####":
                r[0] = nrows
                nrows += 1

        self.twodx.update_rows()
        self.twodx.update_tlks()

        for r in self.twodx.rows:
            highest = max(highest, int(r[0]))

        if highest > 0 and highest >= self.twoda.rows():
            self.twoda.pad((highest - self.twoda.rows()) + 1)

        for c in self.twodx.columns[1:]:
            self.twoda.add_column(c)

        for r in self.twodx.rows:
            row = int(r[0])
            for c in self.twodx.columns[1:]:
                col = self.twodx.column_index(c)
                new = r[col]
                if new == '####':
                    continue
                if (self.default
                   and row < self.default.rows()
                   and self.default.column_index(c) != ctypes.c_uint64(-1).value):
                    orig = self.default.get(row, c)
                    cur = self.twoda.get(row, c)

                    if orig == cur:
                        self.twoda.set(row, c, new)
                else:
                    self.twoda.set(row, c, new)


def safe_mkdir(path):
    if not os.path.exists(path):
        os.makedirs(path)


def get_mergees(base, input_dir):
    matches = []
    pattern = re.compile(rf'{base}_(\d{{2}})\.2dx')

    for root, dirnames, filenames in os.walk(input_dir):
        for filename in filenames:
            if os.path.basename(filename) == f"{base}.2dx" or pattern.match(filename):
                matches.append(os.path.join(root, filename))

    return sorted(matches)


def convert_2das(args):
    """Converts between 2DX and XLSX formats based on file extension and header."""

    for file_path in args.files:
        file_path = Path(file_path)

        if file_path.suffix == ".2dx":
            with open(file_path, 'r') as f:
                content = f.read()
            twodx = TwoDX(content)

            xlsx_path = file_path.with_suffix(".xlsx")
            twodx.to_excel(xlsx_path)

        elif file_path.suffix == ".2da":
            with open(file_path, 'r') as f:
                content = f.read()
            twoda = rollnw.TwoDA.from_string(content)

            xlsx_path = file_path.with_suffix(".xlsx")
            twoda_to_excel(twoda, xlsx_path)

        elif file_path.suffix == ".xlsx":
            workbook = load_workbook(filename=file_path)
            sheet = workbook.active

            header = sheet.cell(row=1, column=1).value
            if header == "2DX V2.1":
                twodx = TwoDX("")
                twodx.from_excel(sheet)
                twodx_path = file_path.with_suffix(".2dx")

                with open(twodx_path, 'w') as f:
                    f.write(str(twodx) + "\n")

                print(f"Converted {file_path} to {twodx_path}")
            elif header == "2DA V2.0":
                twoda = twoda_from_excel(sheet)
                twoda_path = file_path.with_suffix(".2da")

                with open(twoda_path, 'w') as f:
                    f.write(str(twoda))
        else:
            print(f"Unsupported file extension for {file_path}")


def merge_2dx_files(args):
    safe_mkdir(args.output)

    files = []
    if os.name == 'nt':
        for f in args.files:
            files += glob.glob(f)
    else:
        files = args.files

    for file in files:
        start_time = time.time()
        basef = os.path.basename(file)
        base = os.path.splitext(basef)[0]
        twodxs = get_mergees(base, args.input)
        twoda = rollnw.TwoDA(file)

        default = None
        if not args.force:
            with importlib.resources.path("arclight.data", "2dasource.zip") as zip_path:
                with zipfile.ZipFile(zip_path, 'r') as zf:
                    defcont = zf.read(basef)
                    default = rollnw.TwoDA.from_string(defcont)

        merged = False
        for twodx in twodxs:
            with open(twodx, 'r') as f:
                x = TwoDX(f.read())
                merger = TwoDXMerger(twoda, x, default)
                merger.merge()
                merged = True

        if merged:
            with open(os.path.join(args.output, basef), 'w') as f2:
                f2.write(str(twoda))

        end = (time.time() - start_time) * 1000
        print(f"Processed '{file}' in {end:.3f}ms")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('-v', '--version', action='version', version='0.5')

    subparsers = parser.add_subparsers(dest="command")

    merge_parser = subparsers.add_parser(
        "merge", help="Merge and manage 2dx files")
    merge_parser.add_argument(
        '-o', '--output', help='Output directory.', default='merged')
    merge_parser.add_argument(
        '--force', help='Force merges non-default row entries.', action='store_true')
    merge_parser.add_argument(
        'input', help='Directory containing 2dx files to be merged.')
    merge_parser.add_argument('files', help='2da file(s).', nargs='+')

    convert_parser = subparsers.add_parser(
        "convert", help="Converts 2DA/2DX files to/from Excel xlsx")
    convert_parser.add_argument('files', help='File(s) to convert.', nargs='+')

    args = parser.parse_args()

    if args.command == "merge":
        merge_2dx_files(args)
    elif args.command == "convert":
        convert_2das(args)
    else:
        parser.print_help()


if __name__ == "__main__":
    main()
